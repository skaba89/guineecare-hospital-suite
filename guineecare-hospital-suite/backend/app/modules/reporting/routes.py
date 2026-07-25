from datetime import datetime
from app.core.datetime import utcnow

from fastapi import APIRouter, Depends, HTTPException
from sqlalchemy import func
from sqlalchemy.orm import Session

from app.core.pagination import PaginationParams, paginate
from app.core.tenant import tenant_query, enforce_facility_access
from app.db.session import get_db
from app.modules.activity.service import record_activity
from app.modules.reporting.models import EpidemicAlert, HealthStatistic, NationalReport
from app.modules.reporting.schemas import (
    EpidemicAlertCreate,
    HealthStatisticCreate,
    NationalReportCreate,
)
from app.modules.rbac.dependencies import require_permission
from app.modules.users.models import User

router = APIRouter(prefix="/reporting", tags=["reporting"])


# ── National Reports ──────────────────────────────────────────────────

@router.get("/national-reports")
def list_national_reports(
    facility_id: str | None = None,
    status: str | None = None,
    report_type: str | None = None,
    pagination: PaginationParams = Depends(),
    db: Session = Depends(get_db),
    current_user: User = Depends(require_permission("reporting.read")),
):
    query = tenant_query(db, NationalReport, current_user)
    if facility_id:
        query = query.filter(NationalReport.facility_id == facility_id)
    if status:
        query = query.filter(NationalReport.status == status)
    if report_type:
        query = query.filter(NationalReport.report_type == report_type)
    query = query.order_by(NationalReport.created_at.desc())
    return paginate(query, pagination)


@router.post("/national-reports")
def create_national_report(
    payload: NationalReportCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_permission("reporting.manage")),
):
    row = NationalReport(**payload.model_dump(exclude_none=True))
    if not row.facility_id:
        row.facility_id = current_user.facility_id
    enforce_facility_access(current_user, row.facility_id)
    db.add(row)
    db.flush()
    record_activity(
        db=db,
        actor_id=current_user.id,
        action_name="reporting.national_report_created",
        entity_type="national_report",
        entity_id=row.id,
        level="IMPORTANT",
    )
    db.commit()
    db.refresh(row)
    return {"data": row, "message": "national report created"}


@router.post("/national-reports/{report_id}/submit")
def submit_national_report(
    report_id: str,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_permission("reporting.manage")),
):
    report = db.query(NationalReport).filter(NationalReport.id == report_id).first()
    if not report:
        raise HTTPException(status_code=404, detail="National report not found")
    enforce_facility_access(current_user, report.facility_id)
    if report.status != "DRAFT":
        raise HTTPException(status_code=409, detail="Only DRAFT reports can be submitted")

    report.status = "SUBMITTED"
    report.submitted_at = utcnow()
    report.submitted_by = current_user.id

    record_activity(
        db=db,
        actor_id=current_user.id,
        action_name="reporting.national_report_submitted",
        entity_type="national_report",
        entity_id=report.id,
        level="IMPORTANT",
    )
    db.commit()
    db.refresh(report)
    return {"data": report, "message": "national report submitted"}


@router.post("/national-reports/{report_id}/validate")
def validate_national_report(
    report_id: str,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_permission("reporting.manage")),
):
    report = db.query(NationalReport).filter(NationalReport.id == report_id).first()
    if not report:
        raise HTTPException(status_code=404, detail="National report not found")
    enforce_facility_access(current_user, report.facility_id)
    if report.status != "SUBMITTED":
        raise HTTPException(status_code=409, detail="Only SUBMITTED reports can be validated")

    report.status = "VALIDATED"
    report.validated_at = utcnow()
    report.validated_by = current_user.id

    record_activity(
        db=db,
        actor_id=current_user.id,
        action_name="reporting.national_report_validated",
        entity_type="national_report",
        entity_id=report.id,
        level="IMPORTANT",
    )
    db.commit()
    db.refresh(report)
    return {"data": report, "message": "national report validated"}


@router.post("/national-reports/{report_id}/reject")
def reject_national_report(
    report_id: str,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_permission("reporting.manage")),
):
    report = db.query(NationalReport).filter(NationalReport.id == report_id).first()
    if not report:
        raise HTTPException(status_code=404, detail="National report not found")
    enforce_facility_access(current_user, report.facility_id)
    if report.status != "SUBMITTED":
        raise HTTPException(status_code=409, detail="Only SUBMITTED reports can be rejected")

    report.status = "REJECTED"

    record_activity(
        db=db,
        actor_id=current_user.id,
        action_name="reporting.national_report_rejected",
        entity_type="national_report",
        entity_id=report.id,
        level="IMPORTANT",
    )
    db.commit()
    db.refresh(report)
    return {"data": report, "message": "national report rejected"}


# ── Epidemic Alerts ───────────────────────────────────────────────────

@router.get("/epidemic-alerts")
def list_epidemic_alerts(
    facility_id: str | None = None,
    status: str | None = None,
    alert_level: str | None = None,
    pagination: PaginationParams = Depends(),
    db: Session = Depends(get_db),
    current_user: User = Depends(require_permission("reporting.read")),
):
    query = tenant_query(db, EpidemicAlert, current_user)
    if facility_id:
        query = query.filter(EpidemicAlert.facility_id == facility_id)
    if status:
        query = query.filter(EpidemicAlert.status == status)
    if alert_level:
        query = query.filter(EpidemicAlert.alert_level == alert_level)
    query = query.order_by(EpidemicAlert.created_at.desc())
    return paginate(query, pagination)


@router.post("/epidemic-alerts")
def create_epidemic_alert(
    payload: EpidemicAlertCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_permission("reporting.manage")),
):
    row = EpidemicAlert(**payload.model_dump(exclude_none=True))
    if not row.facility_id:
        row.facility_id = current_user.facility_id
    enforce_facility_access(current_user, row.facility_id)
    if not row.reported_by:
        row.reported_by = current_user.id
    db.add(row)
    db.flush()
    record_activity(
        db=db,
        actor_id=current_user.id,
        action_name="reporting.epidemic_alert_created",
        entity_type="epidemic_alert",
        entity_id=row.id,
        level="IMPORTANT",
    )
    db.commit()
    db.refresh(row)
    return {"data": row, "message": "epidemic alert created"}


@router.post("/epidemic-alerts/{alert_id}/close")
def close_epidemic_alert(
    alert_id: str,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_permission("reporting.manage")),
):
    alert = db.query(EpidemicAlert).filter(EpidemicAlert.id == alert_id).first()
    if not alert:
        raise HTTPException(status_code=404, detail="Epidemic alert not found")
    enforce_facility_access(current_user, alert.facility_id)
    if alert.status == "CLOSED":
        raise HTTPException(status_code=409, detail="Alert is already closed")

    alert.status = "CLOSED"
    alert.closed_at = utcnow()

    record_activity(
        db=db,
        actor_id=current_user.id,
        action_name="reporting.epidemic_alert_closed",
        entity_type="epidemic_alert",
        entity_id=alert.id,
        level="IMPORTANT",
    )
    db.commit()
    db.refresh(alert)
    return {"data": alert, "message": "epidemic alert closed"}


# ── Health Statistics ─────────────────────────────────────────────────

@router.get("/statistics")
def list_health_statistics(
    facility_id: str | None = None,
    category: str | None = None,
    period_start: datetime | None = None,
    period_end: datetime | None = None,
    pagination: PaginationParams = Depends(),
    db: Session = Depends(get_db),
    current_user: User = Depends(require_permission("reporting.read")),
):
    query = tenant_query(db, HealthStatistic, current_user)
    if facility_id:
        query = query.filter(HealthStatistic.facility_id == facility_id)
    if category:
        query = query.filter(HealthStatistic.category == category)
    if period_start:
        query = query.filter(HealthStatistic.period_start >= period_start)
    if period_end:
        query = query.filter(HealthStatistic.period_end <= period_end)
    query = query.order_by(HealthStatistic.created_at.desc())
    return paginate(query, pagination)


@router.post("/statistics")
def create_health_statistic(
    payload: HealthStatisticCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_permission("reporting.manage")),
):
    row = HealthStatistic(**payload.model_dump(exclude_none=True))
    if not row.facility_id:
        row.facility_id = current_user.facility_id
    enforce_facility_access(current_user, row.facility_id)
    db.add(row)
    db.commit()
    db.refresh(row)
    return {"data": row, "message": "health statistic created"}


# ── Dashboard ─────────────────────────────────────────────────────────

@router.get("/dashboard")
def get_dashboard(
    facility_id: str | None = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_permission("reporting.read")),
):
    report_q = tenant_query(db, NationalReport, current_user)
    alert_q = tenant_query(db, EpidemicAlert, current_user)
    stat_q = tenant_query(db, HealthStatistic, current_user)

    if facility_id:
        enforce_facility_access(current_user, facility_id)
        report_q = report_q.filter(NationalReport.facility_id == facility_id)
        alert_q = alert_q.filter(EpidemicAlert.facility_id == facility_id)
        stat_q = stat_q.filter(HealthStatistic.facility_id == facility_id)

    total_reports = report_q.count()
    draft_reports = report_q.filter(NationalReport.status == "DRAFT").count()
    submitted_reports = report_q.filter(NationalReport.status == "SUBMITTED").count()
    validated_reports = report_q.filter(NationalReport.status == "VALIDATED").count()

    active_alerts = alert_q.filter(EpidemicAlert.status == "ACTIVE").count()
    total_alerts = alert_q.count()

    total_statistics = stat_q.count()

    return {
        "data": {
            "reports": {
                "total": total_reports,
                "draft": draft_reports,
                "submitted": submitted_reports,
                "validated": validated_reports,
            },
            "alerts": {
                "active": active_alerts,
                "total": total_alerts,
            },
            "statistics": {
                "total": total_statistics,
            },
        },
        "message": "reporting dashboard",
    }


# ============================================================================
# v2.5.0 — Phase 5 : Pilotage national et reporting santé Guinée
# ============================================================================

from app.modules.reporting.national_service import (
    compute_national_dashboard,
    compute_facility_breakdown,
    compute_geographic_distribution,
    export_dhis2_dataset,
)


@router.get("/national")
def national_dashboard(
    region: str | None = None,
    prefecture: str | None = None,
    commune: str | None = None,
    facility_id: str | None = None,
    period: str | None = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_permission("reporting.read")),
):
    """Tableau de bord national agrégé — indicateurs sanitaires multi-établissements.

    SÉCURITÉ (v2.5.0) :
    - Permission `reporting.read` requise (SUPER_ADMIN, ADMIN, DOCTOR).
    - `tenant_query` filtre par facility_id pour les non-SUPER_ADMIN.
    - AUCUNE donnée patient n'est retournée — uniquement des agrégats
      anonymisés (comptages, sommes, moyennes).

    Filtres géographiques (tous optionnels) :
    - region : ex "Conakry", "Kankan", "Kindia"
    - prefecture : ex "Conakry", "Dubreka"
    - commune : ex "Kaloum", "Dixinn"
    - facility_id : UUID d'un établissement spécifique

    Filtre période :
    - period : "2026" (année), "202603" (mois), "2026Q1" (trimestre)
    """
    return compute_national_dashboard(
        db, current_user,
        region=region, prefecture=prefecture, commune=commune,
        facility_id=facility_id, period=period,
    )


@router.get("/facility-breakdown")
def facility_breakdown(
    region: str | None = None,
    prefecture: str | None = None,
    commune: str | None = None,
    period: str | None = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_permission("reporting.read")),
):
    """Activité par établissement — tableau détaillé pour le pilotage national.

    Retourne pour chaque établissement du périmètre :
    patients_count, admissions_count, emergencies_count, lab_orders_count,
    revenue_gnf, outstanding_gnf.

    Trié par activité décroissante.
    """
    breakdown = compute_facility_breakdown(
        db, current_user,
        region=region, prefecture=prefecture, commune=commune, period=period,
    )
    return {
        "data": breakdown,
        "total_facilities": len(breakdown),
        "message": "facility breakdown",
    }


@router.get("/geo-distribution")
def geographic_distribution(
    level: str = "region",  # region | prefecture | commune
    db: Session = Depends(get_db),
    current_user: User = Depends(require_permission("reporting.read")),
):
    """Répartition géographique des établissements et patients.

    level : "region" (défaut) | "prefecture" | "commune"
    """
    if level not in ("region", "prefecture", "commune"):
        raise HTTPException(status_code=422, detail="level doit être region, prefecture ou commune")
    distribution = compute_geographic_distribution(db, current_user, level=level)
    return {
        "data": distribution,
        "level": level,
        "total_zones": len(distribution),
        "message": f"geographic distribution by {level}",
    }


@router.get("/dhis2/{period}")
def dhis2_dataset(
    period: str,
    dataset: str = "SNIS_MENSUEL",
    region: str | None = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_permission("reporting.read")),
):
    """Génère un dataset DHIS2-compatible pour la période donnée.

    Usage : préparation à l'envoi vers une instance DHIS2 nationale.

    Format DHIS2 dataValueSets :
    {
      "dataSet": "SNIS_MENSUEL",
      "period": "202603",
      "orgUnits": ["CHU-DONKA", "CHU-IGNACE-DEEN", ...],
      "dataValues": [
        {"dataElement": "TOTAL_ADMISSIONS", "orgUnit": "CHU-DONKA", "period": "202603", "value": "1234"},
        ...
      ]
    }

    DataElements générés :
    - TOTAL_ADMISSIONS
    - TOTAL_EMERGENCIES
    - TOTAL_DELIVERIES
    - TOTAL_LAB_ORDERS
    - TOTAL_REVENUE_GNF
    """
    if len(period) < 4:
        raise HTTPException(status_code=422, detail="period invalide (format YYYY, YYYYMM ou YYYYQn)")
    return export_dhis2_dataset(
        db, current_user,
        period=period, dataset=dataset, region=region,
    )


@router.get("/export/xlsx")
def export_national_xlsx(
    region: str | None = None,
    prefecture: str | None = None,
    commune: str | None = None,
    period: str | None = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_permission("reporting.read")),
):
    """Export Excel du tableau de bord national + breakdown par établissement.

    Génère un fichier .xlsx avec 2 feuilles :
    1. "Tableau de bord" — indicateurs agrégés + répartition par région
    2. "Par établissement" — activité détaillée par établissement

    Sécurité (v2.5.0) :
    - Permission reporting.read requise
    - Audit log de l'export (action="reporting.export.xlsx")
    - Données anonymisées (aucune donnée patient)
    """
    from fastapi import Response
    from app.modules.audit.service import audit_log
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from io import BytesIO

    # Calculer les données
    dashboard = compute_national_dashboard(
        db, current_user,
        region=region, prefecture=prefecture, commune=commune,
        facility_id=None, period=period,
    )
    breakdown = compute_facility_breakdown(
        db, current_user,
        region=region, prefecture=prefecture, commune=commune, period=period,
    )

    # Audit log
    audit_log(
        db=db,
        action="reporting.export.xlsx",
        user=current_user,
        resource_type="report",
        request=None,  # pas d'objet Request ici — l'audit_log gère None
        status_code=200,
        payload={
            "format": "xlsx",
            "filters": dashboard["filters"],
            "facilities_count": dashboard["facilities_count"],
        },
    )

    # Générer le workbook
    wb = Workbook()

    # Style header
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="0F6B3E", end_color="0F6B3E", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # ── Feuille 1 : Tableau de bord ──
    ws1 = wb.active
    ws1.title = "Tableau de bord"
    ws1["A1"] = "GuinéeCare — Tableau de bord national"
    ws1["A1"].font = Font(bold=True, size=14, color="0F6B3E")
    ws1["A2"] = f"Période : {period or 'Toutes'}"
    ws1["A3"] = f"Filtres : region={region or '-'}, prefecture={prefecture or '-'}, commune={commune or '-'}"
    ws1["A4"] = f"Généré le : {utcnow().isoformat()}"

    ws1["A6"] = "Indicateur"
    ws1["B6"] = "Valeur"
    for col in ["A6", "B6"]:
        ws1[col].font = header_font
        ws1[col].fill = header_fill
        ws1[col].alignment = header_align

    row = 7
    for key, label in [
        ("facilities_count", "Établissements dans le périmètre"),
        ("total_patients", "Total patients"),
        ("total_admissions", "Total admissions"),
        ("active_admissions", "Admissions actives"),
        ("total_consultations", "Total consultations"),
        ("total_emergencies", "Total urgences"),
        ("avg_emergency_wait_min", "Temps d'attente urgences (min)"),
        ("active_stays", "Hospitalisations actives"),
        ("total_beds", "Total lits"),
        ("occupied_beds", "Lits occupés"),
        ("available_beds", "Lits disponibles"),
        ("bed_occupancy_rate", "Taux d'occupation (%)"),
        ("total_pregnancies", "Grossesses suivies"),
        ("total_deliveries", "Accouchements"),
        ("total_products", "Produits pharmacie"),
        ("total_stock_value_gnf", "Valeur stock (GNF)"),
        ("low_stock_count", "Ruptures de stock"),
        ("total_lab_orders", "Demandes laboratoire"),
        ("validated_lab_orders", "Résultats validés"),
        ("pending_lab_orders", "Résultats en attente"),
        ("total_invoices", "Total factures"),
        ("paid_invoices", "Factures payées"),
        ("unpaid_invoices", "Factures impayées"),
        ("total_revenue_gnf", "Recettes (GNF)"),
        ("total_outstanding_gnf", "Créances (GNF)"),
    ]:
        ws1[f"A{row}"] = label
        value = dashboard["indicators"].get(key, dashboard.get(key))
        ws1[f"B{row}"] = value
        row += 1

    # Répartition par région
    row += 2
    ws1[f"A{row}"] = "Répartition par région"
    ws1[f"A{row}"].font = Font(bold=True, size=12)
    row += 1
    ws1[f"A{row}"] = "Région"
    ws1[f"B{row}"] = "Établissements"
    ws1[f"C{row}"] = "Patients"
    for col in [f"A{row}", f"B{row}", f"C{row}"]:
        ws1[col].font = header_font
        ws1[col].fill = header_fill
    row += 1
    for r in dashboard["by_region"]:
        ws1[f"A{row}"] = r["region"]
        ws1[f"B{row}"] = r["facilities_count"]
        ws1[f"C{row}"] = r["patients_count"]
        row += 1

    # Largeurs colonnes
    ws1.column_dimensions["A"].width = 35
    ws1.column_dimensions["B"].width = 20
    ws1.column_dimensions["C"].width = 15

    # ── Feuille 2 : Par établissement ──
    ws2 = wb.create_sheet("Par établissement")
    headers = ["Établissement", "Code", "Région", "Préfecture", "Commune", "Catégorie",
               "Patients", "Admissions", "Urgences", "Labos", "Recettes GNF", "Créances GNF"]
    for col_idx, h in enumerate(headers, 1):
        cell = ws2.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    for row_idx, b in enumerate(breakdown, 2):
        ws2.cell(row=row_idx, column=1, value=b["name"])
        ws2.cell(row=row_idx, column=2, value=b["code"])
        ws2.cell(row=row_idx, column=3, value=b["region"])
        ws2.cell(row=row_idx, column=4, value=b["prefecture"])
        ws2.cell(row=row_idx, column=5, value=b["commune"])
        ws2.cell(row=row_idx, column=6, value=b["category"])
        ws2.cell(row=row_idx, column=7, value=b["patients_count"])
        ws2.cell(row=row_idx, column=8, value=b["admissions_count"])
        ws2.cell(row=row_idx, column=9, value=b["emergencies_count"])
        ws2.cell(row=row_idx, column=10, value=b["lab_orders_count"])
        ws2.cell(row=row_idx, column=11, value=b["revenue_gnf"])
        ws2.cell(row=row_idx, column=12, value=b["outstanding_gnf"])

    # Largeurs
    for col_letter, width in zip("ABCDEFGHIJKL", [30, 15, 15, 15, 15, 20, 12, 12, 12, 12, 18, 18]):
        ws2.column_dimensions[col_letter].width = width

    # ── Serialize ──
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"guineecare_national_{period or 'all'}_{utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ============================================================================
# v2.9.1 — DHIS2 push effectif
# ============================================================================

@router.post("/dhis2/{period}/push")
def push_dhis2(
    period: str,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_permission("reporting.manage")),
):
    """Pousser un dataset DHIS2 vers l'instance nationale.

    v2.9.1 — Push effectif via POST /api/dataValueSets.

    Configuration via variables d'environnement :
    - DHIS2_URL : URL de l'instance DHIS2 (ex: https://dhis2.sante.gov.gn)
    - DHIS2_USERNAME : nom d'utilisateur DHIS2
    - DHIS2_PASSWORD : mot de passe DHIS2

    Si DHIS2_URL n'est pas configurée, retourne le dataset en mode "dry run".

    Sécurité :
    - permission reporting.manage requise (ADMIN, SUPER_ADMIN)
    - audit_log trace le push
    """
    from app.modules.reporting.national_service import push_dhis2_dataset
    from app.modules.audit.service import audit_log

    result = push_dhis2_dataset(db, current_user, period=period)

    # Audit log
    audit_log(
        db=db,
        action="reporting.dhis2.push",
        user=current_user,
        resource_type="dhis2_dataset",
        request=None,
        status_code=200 if result.get("push_status") == "success" else 200,
        payload={
            "period": period,
            "push_status": result.get("push_status"),
            "total_values": result.get("total_values", 0),
        },
    )

    return result
